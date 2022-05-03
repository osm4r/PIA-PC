import openpyxl

def documentExcel(name,name2,name3):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"] = f"{name}"
    sheet["A2"] = " "
    sheet["B1"] = f"{name2}"
    sheet["B2"] = " "
    sheet["C1"] = f"{name3}"
    sheet["C2"] = " "

    with open(f"{name}.txt", "r", encoding="utf-8") as f1:
        i = 3
        for line in f1:
            sheet[f"A{i}"] = line
            i += 1
        with open(f"{name2}" + ".txt", "r", encoding="utf-8")as f2:
            j = 3
            for line2 in f2:
                sheet[f"B{j}"] = line2
                j += 1
            with open(f"{name3}" + ".txt", "r", encoding="utf-8")as f3:
                k = 3
                for line3 in f3:
                    sheet[f"C{k}"] = line3
                    k += 1
                wb.save(filename="Datos.xlsx")


        # wb2 = openpyxl.Workbook()
        # with open(f"{name2}" + ".txt", "r", encoding="utf-8")as f2:
        #     i = 3
        #     for line in f2:
        #         sheet[f"B{i}"] = line
        #         i += 1
        #     wb2.save(filename="Datos.xlsx")
        #
        # wb3 = openpyxl.Workbook()
        # with open(f"{name3}" + ".txt", "r", encoding="utf-8")as f3:
        #     i = 3
        #     for line in f3:
        #         sheet[f"C{i}"] = line
        #         i += 1
        #     wb3.save(filename=f"Datos.xlsx")
        # wb = openpyxl.Workbook()
        # sheet = wb.active
        # with open(f"{name}.txt", "r") as f1:
        #     i = 1
        #     for line in f1:
        #         sheet[f"A{i}"] = line
        #         i += 1
        #     wb.save(filename=f"{name}.xlsx")
        # wb2 = openpyxl.Workbook()
        # sheet2 = wb2.active
        # with open(name2, "r") as f2:
        #     j = 1
        #     for line in f2:
        #         sheet2[f"A{j}"] = line
        #         j += 1
        #     wb2.save(filename=f"{name2}.xlsx")
        # wb3 = openpyxl.Workbook()
        # sheet3 = wb3.active
        # with open(name3, "r") as f3:
        #     k = 1
        #     for line in f3:
        #         sheet3[f"A{k}"] = line
        #         k += 1
        #     wb3.save(filename=f"{name3}.xlsx")
    # except:
    #     pass

# def documentExcel(name):
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#     sheet["A1"] = "Links"
#     with open(f"{name}.txt","r") as f1:
#         i=2
#         for line in f1:
#             sheet[f"A{i}"] = line
#             i+=1
#         wb.save(filename = f"{name}.xlsx")
# documentExcel("tempNoticias")
# #
# def documentExcel(name,name2,name3):
#     wb = Workbook()
#     ws = wb.active()
#     with open(f"{name}.txt","r") as f1:
#         i=1
#         for line in f1:
#             ws[f"A{i}"] = line
#             i+=1
#         wb.save(f"{name}.xlsx")
#     wb2 = Workbook()
#     ws2 = wb2.active()
#     with open(name2,"r") as f2:
#         j=1
#         for line in f2:
#             ws2[f"A{j}"] = line
#             j+=1
#         wb2.save(f"{name2}.xlsx")
#     wb3 = Workbook()
#     ws3 = wb3.active()
#     with open(name2,"r") as f3:
#         k=1
#         for line in f3:
#             ws3[f"A{k}"] = line
#             k+=1
#         wb.save(f"{name3}.xlsx")
# documentExcel("Datos","Instrucciones_PIA.txt", "Titulares.txt")